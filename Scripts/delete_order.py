from openpyxl import load_workbook
from Scripts.xlsx_to_list import Xlsx_to_list
from Scripts.get_real_index import GetRealIndex

class DeleteOrder:
    def __init__(self, order_id:list, customer_list:list):
        self.order_id = order_id
        self.customer_list = customer_list
    
    def deletar_encomenda(self) -> None:
        arquivo = load_workbook("dados.xlsx")
        planilha = arquivo.active
    
        index = GetRealIndex(self.customer_list, self.order_id).return_index()  

        planilha.delete_rows(index)
        arquivo.save("dados.xlsx")
        self.atualizar_id_tabela()

    def atualizar_id_tabela(self) -> None:
        arquivo = load_workbook("dados.xlsx")
        planilha = arquivo.active

        id = Xlsx_to_list("A").toListNum()

        for i in range(len(id)):
            planilha.cell(row=i+2, column=1).value = i + 1
        
        arquivo.save("dados.xlsx")


