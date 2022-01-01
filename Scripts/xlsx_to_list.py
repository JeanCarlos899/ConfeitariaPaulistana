from openpyxl import load_workbook

class Xlsx_to_list:
    def __init__(self, coluna:str):
        self.coluna = coluna

    def toListNum(self) -> list:
        planilha = load_workbook("dados.xlsx")
        aba_ativa = planilha.active
        lista = []
        for celula in aba_ativa[self.coluna]:
            valores = celula.value
            if valores != None and valores != "" and type(valores) != str:
                    lista.append(valores)
        return list(lista)

    def toListStr(self) -> list:
        planilha = load_workbook("dados.xlsx")
        aba_ativa = planilha.active
        lista = []
        for celula in aba_ativa[self.coluna]:
            valores = celula.value
            if valores == None:
                valores = " "
            lista.append(valores)
        lista.pop(0)
        return list(lista)

