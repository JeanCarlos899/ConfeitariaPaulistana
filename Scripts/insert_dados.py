from Scripts.xlsx_to_list import Xlsx_to_list
from openpyxl import load_workbook, Workbook
from datetime import datetime

while True:
    try:
        teste = load_workbook("dados.xlsx")
        break
    except:
        wb = Workbook()
        ws = wb.active
        ws.title = "Dados"
        ws["A1"] = "ID"
        ws["B1"] = "Nome cliente"
        ws["C1"] = "Data de entrega"
        ws["D1"] = "Hora de entrega"
        ws["E1"] = "Bolo de aniversário"
        ws["F1"] = "Bolo de casamento"
        ws["G1"] = "Salgado mini"
        ws["H1"] = "Salgado normal"
        ws["I1"] = "Valor final"
        ws["J1"] = "Mensagem adicional"
        ws["K1"] = "Status"
        wb.save("dados.xlsx")
        continue

class InsertDados:
    def __init__(self, lista_dados: list):
        self.lista_dados = lista_dados
        self.status = "Pendente"

    def verificar_data(self, horario_entrega: str, data_entrega: str) -> bool:
        if horario_entrega.count(":") == 1:
            if len(horario_entrega.split(":")) == 2:
                
                try:
                    hora_entrada = datetime.strptime(
                        (str(data_entrega) + " " + str(horario_entrega)), "%d/%m/%Y %H:%M"
                        ).strftime("%d/%m/%Y %H:%M")
                    hora_atual = datetime.today().strftime('%d/%m/%Y %H:%M')
                    
                    if hora_entrada >= hora_atual:
                        return True
                except:
                    return False
            else:
                return False
        else:
            return False

    def inserir_dados(self) -> bool:
        dados = load_workbook("dados.xlsx")
        planilha_ativa = dados.active

        try:
            num = max(Xlsx_to_list("A").toListNum()) + 1
        except:
            num = 1

        if self.lista_dados[0] != "":
            if self.verificar_data(self.lista_dados[2], self.lista_dados[1]) == True:
                if int(self.lista_dados[3]) > 0 or int(self.lista_dados[4]) > 0:
                    if (int(self.lista_dados[5]) + int(self.lista_dados[6]) >= 25 
                        or int(self.lista_dados[5]) + int(self.lista_dados[6]) == 0):

                        letras = ["B", "C", "D", "E", "F", "G", "H", "J"]

                        planilha_ativa[f"A{num+1}"] = num
                        for i in range(len(self.lista_dados)):
                            planilha_ativa[letras[i] + str(num+1)] = self.lista_dados[i]
                        planilha_ativa[f"K{num+1}"] = self.status 

                        dados.save("dados.xlsx")
                        return True
                    else:
                        return False
                else:
                    return False
            else:
                return False
        else:
            return False